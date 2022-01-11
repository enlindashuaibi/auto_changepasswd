from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
import openpyxl
#宝德
def changepasswd(url,passwd):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get('http://' + url +'/')
    driver.find_element_by_xpath("//*[@id='body']/div/div[2]/button[3]").click()
    driver.find_element_by_id('proceed-link').click()
    time.sleep(3)
    driver.find_element_by_id('iduserName').send_keys('admin')
    driver.find_element_by_xpath("//*[@id='wrap']/div[2]/div/div[2]/form/div/div[2]/div[2]/input").send_keys('oldpasswd')
    driver.find_element_by_xpath("//*[@id='wrap']/div[2]/div/div[2]/form/div/div[2]/div[3]/button").click()
    time.sleep(10)
    #a = driver.switch_to.alert
    #print(a.text)
    #a.accept()
    #time.sleep(5)
    driver.get('http://' + url +'/main.html#/configUser')
    time.sleep(15)
    driver.find_element_by_xpath("//*[@id='rightDiv']/div[1]/div/table[4]/tbody/tr[1]/td[2]/div").click()
    driver.find_element_by_xpath("//*[@id='rightDiv']/div[1]/div/table[4]/tbody/tr[1]/td[2]/div").click()
    time.sleep(3)
    driver.find_element_by_xpath("//*[@id='idBoxSigGroup']/div[2]/table/tbody/tr[2]/td[2]").click()
    driver.find_element_by_id('idOldPassword').send_keys('oldpasswd')
    driver.find_element_by_id('idPassword').send_keys(passwd)
    driver.find_element_by_id('idPasswordCfm').send_keys(passwd)
    driver.find_element_by_xpath("//*[@id='idBoxSigGroup']/div[3]/button").click()
    b = driver.switch_to.alert
    print(b.text)
    b.accept()
    driver.close()
#浪潮
def changepasswd2(url,passwd):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get('https://' + url)
    driver.find_element_by_xpath("//*[@id='body']/div/div[2]/button[3]").click()
    driver.find_element_by_id('proceed-link').click()
    time.sleep(3)
    driver.find_element_by_xpath("//*[@id='usrName']").click()
    driver.find_element_by_xpath("//*[@id='usrName']").send_keys('ADMIN')
    driver.find_element_by_xpath("//*[@id='pwd']").click()
    driver.find_element_by_xpath("//*[@id='pwd']").send_keys('oldpasswd')
    driver.find_element_by_id('login_word').click()
    time.sleep(3)
    driver.find_element_by_xpath('//*[@id="_submenu"]/li[3]/a/span').click()
    driver.find_element_by_xpath('//*[@id="menu_config_account_overview"]').click()
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="userDataTable"]/table/tbody/tr[2]/td[5]/a/img').click()
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="userLockLink"]').click()
    driver.find_element_by_id('password_pwd').click()
    driver.find_element_by_id('password_pwd').send_keys(passwd)
    driver.find_element_by_id('password_pwd_check').click()
    driver.find_element_by_id('password_pwd_check').send_keys(passwd)
    driver.find_element_by_id('formSubmitBtn').click()
    driver.close()
def main(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb['Sheet1']
    i = 1
    # 填入真实行数
    while (i <= 33):
        # 第G列，邮箱账号
        url = sheet.cell(row=i, column=1).value
        passwd = 'newpasswd'
        if sheet.cell(row=i, column=3).value != '已修改':
            try:
                changepasswd2(url, passwd)
            except Exception as e:
                sheet.cell(row=i, column=3).value = '遇到异常，未修改'
                wb.save(file)
                print(e)
                i = i + 1
                continue
            sheet.cell(row=i, column=3).value = '已修改'
            print(" %s已修改" % (url))
            wb.save(file)
        i = i + 1
    print('All host is Created!')
#add comment
main("test.xlsx")


